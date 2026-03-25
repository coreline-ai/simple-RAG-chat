"""엑셀 이슈 데이터 파서

형식: model_issue_dataset_10000.xlsx
스킬 원칙:
  - 1 row = 1 사건 = 1 chunk (기본)
  - 문제점 분석 텍스트는 통째로 유지 (가능하면)
  - 긴 row는 2차 KSS 분할
  - 날짜는 반드시 metadata로 분리
"""
from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path
from typing import Any

from app.config import settings
from app.services.parsers.base import BaseParser


# 엑셀 컬럼명 → 내부 키 매핑
COLUMN_MAP = {
    "모델 이슈 검토 사항": "title",
    "등록일": "created_at",
    "기본 확인내용": "basic_check",
    "기본 작업내용": "basic_work",
    "업무지시": "instruction",
    "담당자": "assignee",
    "업무시작일": "start_at",
    "완료예정": "due_at",
    "진행(담당자)": "progress",
    "완료일": "completed_at",
    "문제점 분석 내용 (담당자 Comments)": "analysis",
    "상태_도우미": "status",
    "등록월_도우미": "created_month",
}


class ExcelIssueParser(BaseParser):
    """엑셀 이슈 데이터 (.xlsx) 파서"""

    def detect(self, filename: str) -> bool:
        return filename.endswith((".xlsx", ".xls"))

    def parse(self, source: Any) -> list[dict]:
        """엑셀 파일을 파싱하여 청크 리스트 반환

        Args:
            source: 파일 경로(str/Path) 또는 바이너리(bytes/BytesIO)
        """
        import openpyxl

        if isinstance(source, (str, Path)):
            wb = openpyxl.load_workbook(str(source), read_only=True, data_only=True)
        elif isinstance(source, bytes):
            wb = openpyxl.load_workbook(io.BytesIO(source), read_only=True, data_only=True)
        elif isinstance(source, io.BytesIO):
            wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
        else:
            raise ValueError(f"지원하지 않는 소스 타입: {type(source)}")

        # 시트 선택
        sheet_name = settings.excel_sheet_name
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            return []

        # 헤더 매핑
        headers = [str(h).strip() if h else "" for h in rows[0]]
        col_indices = {}
        for i, header in enumerate(headers):
            if header in COLUMN_MAP:
                col_indices[COLUMN_MAP[header]] = i

        # 데이터 파싱
        results = []
        id_prefix = settings.excel_id_prefix

        for row_idx, row in enumerate(rows[1:], start=1):
            row_data = self._extract_row(row, col_indices)
            if not row_data:
                continue

            chunks = self._build_chunks(row_data, row_idx, id_prefix)
            results.extend(chunks)

        return results

    def _extract_row(self, row: tuple, col_indices: dict) -> dict | None:
        """엑셀 행에서 데이터 추출"""
        data = {}
        for key, idx in col_indices.items():
            if idx < len(row):
                data[key] = row[idx]
            else:
                data[key] = None

        # 최소 필수 필드 확인
        if not data.get("title"):
            return None

        return data

    def _build_chunks(self, row: dict, row_idx: int, id_prefix: str) -> list[dict]:
        """Row Chunking: 1 row → 1+ chunks

        스킬 원칙: "Row는 문장이 아니라 사건이다"
        기본적으로 1 row = 1 chunk. row 전체가 너무 길면 2차 KSS 분할.
        """
        doc_id = f"{id_prefix}_{row_idx:06d}"
        metadata = self._build_metadata(row, doc_id)

        # 임베딩 텍스트 조립
        full_text = self._build_embedding_text(row)

        # 2차 분할 판단
        if len(full_text) > settings.excel_row_max_chars and row.get("analysis"):
            return self._split_row(row, full_text, metadata, doc_id)

        # 1 row = 1 chunk
        original = full_text  # 원본으로도 사용
        return [{
            "embedding_text": full_text,
            "original": original,
            "metadata": {**metadata, "split_index": 0},
        }]

    def _build_embedding_text(self, row: dict) -> str:
        """Row를 구조화된 임베딩 텍스트로 조립"""
        parts = []
        parts.append(f"[이슈] {_safe_str(row.get('title'))}")
        parts.append(f"[등록일] {_format_date(row.get('created_at'))}")
        parts.append(f"[기본 확인내용] {_safe_str(row.get('basic_check'))}")
        parts.append(f"[기본 작업내용] {_safe_str(row.get('basic_work'))}")
        parts.append(f"[업무지시] {_safe_str(row.get('instruction'))}")
        parts.append(f"[담당자] {_safe_str(row.get('assignee'))}")
        parts.append(f"[진행] {_safe_str(row.get('progress'))}")

        completed = row.get("completed_at")
        if completed:
            parts.append(f"[완료일] {_format_date(completed)}")

        analysis = row.get("analysis")
        if analysis:
            parts.append(f"[문제 원인 분석 결과] {_safe_str(analysis)}")

        return "\n".join(parts)

    def _split_row(self, row: dict, full_text: str, metadata: dict, doc_id: str) -> list[dict]:
        """2차 분할: 이슈 요약 + 행동 흐름 기반 분석 chunk

        하이브리드 전략:
          1단계: KSS 문장 분리
          2단계: 규칙 기반 라벨링 (discovery/attempt/fix/result 등)
          3단계: 인접 라벨 묶기 (행동 흐름 chunk)
        """
        from app.services.parsers.labeler import build_flow_chunks, split_and_label

        title = _safe_str(row.get("title"))

        # Chunk 1: 이슈 요약
        summary_parts = [
            f"[이슈] {title}",
            f"[등록일] {_format_date(row.get('created_at'))}",
            f"[기본 확인내용] {_safe_str(row.get('basic_check'))}",
            f"[기본 작업내용] {_safe_str(row.get('basic_work'))}",
            f"[업무지시] {_safe_str(row.get('instruction'))}",
            f"[담당자] {_safe_str(row.get('assignee'))}",
        ]
        summary_text = "\n".join(summary_parts)

        chunks = [{
            "embedding_text": summary_text,
            "original": full_text,
            "metadata": {**metadata, "split_index": 0, "flow_name": "이슈 요약"},
        }]

        # Chunk 2+: 문제 원인 분석 (행동 흐름 기반)
        analysis = _safe_str(row.get("analysis"))
        if analysis:
            labeled = split_and_label(analysis)
            flow_chunks = build_flow_chunks(labeled)

            for i, fc in enumerate(flow_chunks):
                analysis_text = f"[이슈] {title}\n[{fc['flow_name']}] {fc['text']}"
                chunks.append({
                    "embedding_text": analysis_text,
                    "original": full_text,
                    "metadata": {
                        **metadata,
                        "split_index": i + 1,
                        "flow_name": fc["flow_name"],
                        "labels": ",".join(fc["labels"]),
                    },
                })

        return chunks

    def _build_metadata(self, row: dict, doc_id: str) -> dict:
        """메타데이터 생성 — 날짜 3형식 + 하위 호환"""
        created_at = row.get("created_at")
        created_iso = _format_date(created_at)
        created_int = _datetime_to_int(created_at)

        return {
            "doc_type": "issue",
            "doc_id": doc_id,
            "title": _safe_str(row.get("title")),
            "assignee": _safe_str(row.get("assignee")),
            "status": _safe_str(row.get("status")),
            # 날짜 3형식
            "created_at_iso": created_iso,
            "created_at_int": created_int,
            "start_at_int": _datetime_to_int(row.get("start_at")),
            "due_at_int": _datetime_to_int(row.get("due_at")),
            "completed_at_int": _datetime_to_int(row.get("completed_at")),
            # 하위 호환: date/date_int (= created_at 기준)
            "date": created_iso,
            "date_int": created_int,
        }


# === 날짜 변환 유틸 ===

def _datetime_to_int(dt: Any) -> int:
    """datetime을 YYYYMMDD int로 변환. None이면 0"""
    if dt is None:
        return 0
    if isinstance(dt, datetime):
        return int(dt.strftime("%Y%m%d"))
    if isinstance(dt, str):
        try:
            return int(dt.replace("-", "").replace("/", "")[:8])
        except (ValueError, IndexError):
            return 0
    return 0


def _format_date(dt: Any) -> str:
    """datetime을 YYYY-MM-DD 문자열로 변환"""
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m-%d")
    return str(dt).strip()


def _safe_str(val: Any) -> str:
    """None-safe 문자열 변환"""
    if val is None:
        return ""
    return str(val).strip()
